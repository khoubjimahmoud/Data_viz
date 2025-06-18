import pandas as pd
from openpyxl import load_workbook

def read_with_style(upl, sheet_name=0, gray_rgb="FF7a7a7a"):
    """
    Read into pandas + openpyxl. Return (df, gray_mask)
    where gray_mask[i,j] == True if that cell is shaded gray.
    """
    df = pd.read_excel(upl, sheet_name=sheet_name, engine="openpyxl")
    wb = load_workbook(upl, data_only=True)
    ws = wb[wb.sheetnames[sheet_name] if isinstance(sheet_name,int) else sheet_name]

    mask = pd.DataFrame(False, index=df.index, columns=df.columns)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        for j, cell in enumerate(row):
            if cell.fill and getattr(cell.fill.fgColor, "rgb", None) == gray_rgb:
                mask.iat[i,j] = True
    return df, mask
